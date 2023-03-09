import fitz # install using: pip install PyMuPDF
import openpyxl
import sys
import re

pdfdocument = sys.argv[1]

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Data"

row_number = 2 # start from row 2 to leave the first row for headers
page_check = 0
with fitz.open(pdfdocument) as doc:

    for page in range(len(doc)):
        
        current_id = ''
        text = doc[page].get_text()
        data_list = text.split("\n")
        # print(data_list)
        if data_list == ['']:
            continue

        for data in range(len(data_list)):
            if not data_list[data]:
                continue

            # check if data starts a new row
            if data_list[data].startswith('ADBG') or \
                data_list[data].startswith('ADFT') or \
                data_list[data].startswith('AD') or \
                data_list[data].startswith('COW') or \
                data_list[data].startswith('GI') or \
                data_list[data].startswith('ADTL') or \
                    \
                    \
                re.match(r'^NMBG.*', data_list[data]) or \
                    \
                    \
                re.match(r'^KB.*', data_list[data]) or \
                    \
                    \
                re.match(r'^SWC.*', data_list[data]):
                # update current ID
                current_id = data_list[data]

                # write ID to first column of new row
                ws.cell(row=row_number, column=1).value = current_id

                # increment row_number for next data row
                row_number += 1

            # check for size data
            elif 'Size: ' in data_list[data] or\
                data_list[data] == "Size:":
                # write size to second column of current row
                if data_list[data] == "Size:":
                    ws.cell(row=row_number-1, column=2).value = data_list[data+1]
                else:
                    ws.cell(row=row_number-1, column=2).value = data_list[data].split(': ')[1]

            # check for price data
            elif 'Price: ' in data_list[data]:
                # write price to third column of current row
                ws.cell(row=row_number-1, column=3).value = data_list[data].split(': ')[1]

            # check for quantity data
            elif 'Qty: ' in data_list[data] or\
                data_list[data] == "Qty:":
                # write quantity to fourth column of current row
                if data_list[data] == "Qty:":
                    ws.cell(row=row_number-1, column=4).value = data_list[data+1]
                else:
                    ws.cell(row=row_number-1, column=4).value = data_list[data].split(': ')[1]

            elif 'DISCONTINUED' in data_list[data]:
                # write quantity to fourth column of current row
                ws.cell(row=row_number-1, column=5).value = data_list[data]

            # check for page data
            elif 'page:' in data_list[data]:                
                if page_check == 0:
                    start_row_num = 2
                    for num in range(start_row_num,row_number):
                        # print(num)
                        ws.cell(row=num, column=7).value = data_list[data].split(':')[1].strip()
                    page_check += 1
                    row_start_tracker = row_number
                else:
                    row_end_tracker = row_number
                    for numb in range(row_start_tracker,row_end_tracker):
                        ws.cell(row=numb, column=7).value = data_list[data].split(':')[1].strip()
                    row_start_tracker = row_number
               
            # write category to fifth column of current row
            ws.cell(row=row_number-1, column=6).value = data_list[0]

# write headers to first row after data has been written
header_row = ['ID', 'Size', 'Price', 'Quantity', 'Status', 'Category', 'Page']
for col, header in enumerate(header_row, start=1):
    ws.cell(row=1, column=col).value = header
            
# save workbook to file
wb.save(pdfdocument+'.xlsx')
